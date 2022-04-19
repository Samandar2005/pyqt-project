import sys
import time
import traceback
from PyQt5.QtWidgets import *
from PyQt5.QtGui import *
from PyQt5.QtCore import *
from models import District, Student
from windows.RegionWindow import RegionWindow
from windows.DistrictWindow import DistrictWindow
from openpyxl import Workbook


class Window(QMainWindow):

    def __init__(self) -> None:
        super().__init__()

        self.initUI()
        self.initActions()
        self.initMenu()
        self.initTable()
        self.fillTable()
        self.msg = QMessageBox()

    def initUI(self):
        btn_add = QPushButton("Add", self)
        btn_add.move(30, 30)
        btn_add.clicked.connect(self.onAdd)

        btn_add = QPushButton("Update", self)
        btn_add.move(130, 30)
        btn_add.clicked.connect(self.onUpdate)

        btn_add = QPushButton("Delete", self)
        btn_add.move(230, 30)
        btn_add.clicked.connect(self.onDel)

        btn_add = QPushButton("Report", self)
        btn_add.move(330, 30)
        btn_add.clicked.connect(self.onRep)

        ql = QLabel("Fam: ", self)
        ql.move(1050, 60)
        ql = QLabel("Ism: ", self)
        ql.move(1050, 90)
        ql = QLabel("Tug yil: ", self)
        ql.move(1050, 120)
        ql = QLabel("Mark: ", self)
        ql.move(1050, 150)
        ql = QLabel("Level: ", self)
        ql.move(1050, 180)
        ql = QLabel("Dist: ", self)
        ql.move(1050, 210)

        self.qle_fam = QLineEdit(self)
        self.qle_fam.move(1150, 60)
        self.qle_ism = QLineEdit(self)
        self.qle_ism.move(1150, 90)
        self.qle_tug = QLineEdit(self)
        self.qle_tug.move(1150, 120)
        self.qle_mark = QLineEdit(self)
        self.qle_mark.move(1150, 150)
        self.qle_lev = QLineEdit(self)
        self.qle_lev.move(1150, 180)
        self.cbb_dist = QComboBox(self)
        self.cbb_dist.move(1150, 210)
        for item in District.objects():
            self.cbb_dist.addItem(item.name, item.id)

    def onRep(self):
        wb = Workbook()
        try:
            # grab the active worksheet
            ws = wb.active
            ws[f'A{1}'] = "Familiya"
            ws[f'B{1}'] = "Ism"
            ws[f'C{1}'] = "Tug'ulgan yili"
            ws[f'D{1}'] = "Mark"
            ws[f'E{1}'] = "Level"
            ws[f'F{1}'] = "Viloyati"
            ws[f'G{1}'] = "Tumani"

            for sel_row in range(self.table.rowCount()):
                fam = self.table.item(sel_row, 1).text()
                ism = self.table.item(sel_row, 2).text()
                tug = int(self.table.item(sel_row, 3).text())
                mark = int(self.table.item(sel_row, 4).text())
                level = int(self.table.item(sel_row, 5).text())
                reg_name = self.table.item(sel_row, 6).text()
                dist_name = self.table.item(sel_row, 8).text()

                ws[f'A{sel_row + 2}'] = fam
                ws[f'B{sel_row + 2}'] = ism
                ws[f'C{sel_row + 2}'] = tug
                ws[f'D{sel_row + 2}'] = mark
                ws[f'E{sel_row + 2}'] = level
                ws[f'F{sel_row + 2}'] = reg_name
                ws[f'G{sel_row + 2}'] = dist_name
            # Save the file
            wb.save("sample.xlsx")
            wb.close()

        except Exception as exp:
            self.msg.setIcon(QMessageBox.Critical)
            self.msg.setWindowTitle("Xatolik")
            self.msg.setText(str(exp))
            self.msg.show()
            wb.close()
            traceback.print_exc()

    def onAdd(self):
        try:
            fam = self.qle_fam.text()
            ism = self.qle_ism.text()
            tug = int(self.qle_tug.text())
            mark = int(self.qle_mark.text())
            level = int(self.qle_lev.text())
            distId = self.cbb_dist.currentData()

            student = Student(fam, ism, tug, mark, level, distId)
            student.save()

            dist = student.district
            reg = dist.region
            row_count = self.table.rowCount()
            self.table.setRowCount(row_count + 1)
            self.table.setItem(row_count, 0,
                               QTableWidgetItem(str(student.id)))
            self.table.setItem(row_count, 1,
                               QTableWidgetItem(student.fam))
            self.table.setItem(row_count, 2,
                               QTableWidgetItem(student.ism))
            self.table.setItem(row_count, 3,
                               QTableWidgetItem(str(student.tug_yil)))
            self.table.setItem(row_count, 4,
                               QTableWidgetItem(str(student.mark)))
            self.table.setItem(row_count, 5,
                               QTableWidgetItem(str(student.level)))

            self.table.setItem(row_count, 6,
                               QTableWidgetItem(str(reg.name)))
            self.table.setItem(row_count, 7,
                               QTableWidgetItem(str(dist.id)))
            self.table.setItem(row_count, 8,
                               QTableWidgetItem(dist.name))
            self.msg.setIcon(QMessageBox.Information)
            self.msg.setWindowTitle("Bajarildi")
            self.msg.setText("Talaba saqlandi....")
            self.msg.show()
        except Exception as exp:
            self.msg.setIcon(QMessageBox.Critical)
            self.msg.setWindowTitle("Xatolik")
            self.msg.setText(str(exp))
            self.msg.show()

            traceback.print_exc()

    def onUpdate(self):
        try:
            id = int(self.table.item(self.sel_row, 0).text())
            fam = self.qle_fam.text()
            ism = self.qle_ism.text()
            tug = int(self.qle_tug.text())
            mark = int(self.qle_mark.text())
            level = int(self.qle_lev.text())
            distId = self.cbb_dist.currentData()

            student = Student(fam, ism, tug, mark, level, distId, id)
            student.save()

            dist = student.district
            reg = dist.region
            row_count = self.table.currentRow()
            self.table.setItem(row_count, 0,
                               QTableWidgetItem(str(student.id)))
            self.table.setItem(row_count, 1,
                               QTableWidgetItem(student.fam))
            self.table.setItem(row_count, 2,
                               QTableWidgetItem(student.ism))
            self.table.setItem(row_count, 3,
                               QTableWidgetItem(str(student.tug_yil)))
            self.table.setItem(row_count, 4,
                               QTableWidgetItem(str(student.mark)))
            self.table.setItem(row_count, 5,
                               QTableWidgetItem(str(student.level)))

            self.table.setItem(row_count, 6,
                               QTableWidgetItem(str(reg.name)))
            self.table.setItem(row_count, 7,
                               QTableWidgetItem(str(dist.id)))
            self.table.setItem(row_count, 8,
                               QTableWidgetItem(dist.name))
            self.msg.setIcon(QMessageBox.Information)
            self.msg.setWindowTitle("Bajarildi")
            self.msg.setText("Talaba saqlandi....")
            self.msg.show()
        except Exception as exp:
            self.msg.setIcon(QMessageBox.Critical)
            self.msg.setWindowTitle("Xatolik")
            self.msg.setText(str(exp))
            self.msg.show()

            traceback.print_exc()

    def onDel(self):
        try:
            id = int(self.table.item(self.sel_row, 0).text())
            self.qle_fam.setText('')
            self.qle_ism.setText('')
            self.qle_tug.setText('')
            self.qle_mark.setText('')
            self.qle_lev.setText('')

            row_count = self.table.currentRow()
            self.table.removeRow(row_count)

            self.msg.setIcon(QMessageBox.Information)
            self.msg.setWindowTitle("Bajarildi")
            self.msg.setText("Talaba o'chirildi....")
            self.msg.show()
        except Exception as exp:
            self.msg.setIcon(QMessageBox.Critical)
            self.msg.setWindowTitle("Xatolik")
            self.msg.setText(str(exp))
            self.msg.show()

            traceback.print_exc()

    def initTable(self):
        self.table = QTableWidget(self)  # Создаём таблицу
        self.table.move(30, 60)
        self.table.setMinimumSize(1000, 600)
        self.table.setColumnCount(9)     # Устанавливаем три колонки
        self.table.setSelectionBehavior(QTableWidget.SelectRows)

        # Устанавливаем заголовки таблицы
        self.table.setHorizontalHeaderLabels(
            ['Id', "Familiyasi", "Ismi", "Tug'ilgan yili", "Bahosi", "Kursi", "Viloyati", "tuman id",

             "Tumani"])

        self.table.hideColumn(0)
        self.table.hideColumn(7)
        self.table.clicked.connect(self.onClicked)

    def initActions(self):
        self.newAction = QAction("&New...", self)
        self.newAction.triggered.connect(self.onnewAction)
        self.openAction = QAction("&Open...", self)
        self.saveAction = QAction("&Save", self)
        self.exitAction = QAction("&Exit", self)

        self.regionAction = QAction("&Regions", self)
        self.regionAction.triggered.connect(self.onRegionWindow)
        self.districtAction = QAction("&Districts", self)
        self.districtAction.triggered.connect(self.onDistrictWindow)

    def fillTable(self):

        for student in Student.objects():
            dist = student.district
            reg = dist.region
            row_count = self.table.rowCount()
            self.table.setRowCount(row_count + 1)
            self.table.setItem(row_count, 0,
                               QTableWidgetItem(str(student.id)))
            self.table.setItem(row_count, 1,
                               QTableWidgetItem(student.fam))
            self.table.setItem(row_count, 2,
                               QTableWidgetItem(student.ism))
            self.table.setItem(row_count, 3,
                               QTableWidgetItem(str(student.tug_yil)))
            self.table.setItem(row_count, 4,
                               QTableWidgetItem(str(student.mark)))
            self.table.setItem(row_count, 5,
                               QTableWidgetItem(str(student.level)))

            self.table.setItem(row_count, 6,
                               QTableWidgetItem(str(reg.name)))
            self.table.setItem(row_count, 7,
                               QTableWidgetItem(str(dist.id)))
            self.table.setItem(row_count, 8,
                               QTableWidgetItem(dist.name))
            # делаем ресайз колонок по содержимому
        self.table.resizeColumnsToContents()

    def onClicked(self):
        try:
            self.sel_row = self.table.currentRow()
            fam = self.table.item(self.sel_row, 1).text()
            self.qle_fam.setText(fam)
            ism = self.table.item(self.sel_row, 2).text()
            self.qle_ism.setText(ism)
            tug = int(self.table.item(self.sel_row, 3).text())
            self.qle_tug.setText(str(tug))
            mark = int(self.table.item(self.sel_row, 4).text())
            self.qle_mark.setText(str(mark))
            level = int(self.table.item(self.sel_row, 5).text())
            self.qle_lev.setText(str(level))
            distId = int(self.table.item(self.sel_row, 7).text())
            for i in range(self.cbb_dist.count()):
                if self.cbb_dist.itemData(i) == distId:
                    self.cbb_dist.setCurrentIndex(i)
                    break
        except Exception as exp:
            self.msg.setIcon(QMessageBox.Critical)
            self.msg.setWindowTitle("Xatolik")
            self.msg.setText(str(exp))
            self.msg.show()

            traceback.print_exc()

    def onnewAction(self):
        pass

    def initMenu(self):
        menuBar = self.menuBar()
        self.setMenuBar(menuBar)

        fileMenu = menuBar.addMenu("&File")
        fileMenu.addAction(self.newAction)
        fileMenu.addAction(self.openAction)
        fileMenu.addAction(self.saveAction)
        fileMenu.addAction(self.exitAction)

        servicesMenu = menuBar.addMenu("&Services")
        servicesMenu.addAction(self.regionAction)
        servicesMenu.addAction(self.districtAction)

        helpMenu = menuBar.addMenu("&Help")

    def onRegionWindow(self):
        self.regw = RegionWindow()
        self.regw.show()

    def onDistrictWindow(self):
        self.distw = DistrictWindow()
        self.distw.show()


app = QApplication(sys.argv)

w = Window()
w.showMaximized()

app.exec()
